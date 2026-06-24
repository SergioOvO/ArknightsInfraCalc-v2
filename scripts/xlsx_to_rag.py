"""
Convert 公孙长乐工具人表 xlsx → RAG-ready markdown.
Each column section becomes a self-contained paragraph with context.
"""
import sys
import openpyxl


def parse_section(ws, name: str, col_start: int, col_end: int, start_row: int = 4):
    """Extract all rows of a column section, merging multi-row entries."""
    entries = []
    current = []
    current_title = ""

    for row_idx in range(start_row, ws.max_row + 1):
        cells = []
        for col_idx in range(col_start, col_end + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is not None:
                cells.append(str(v).strip())

        if not cells:
            # Empty row → flush current entry
            if current and current_title:
                entries.append((current_title, "; ".join(current)))
            current = []
            current_title = ""
            continue

        text = "; ".join(cells)

        # First row of section = header/title
        if row_idx == start_row:
            if text:
                current_title = text
            continue

        # Check if this row starts a new entry (first cell has content)
        first_cell = str(ws.cell(row=row_idx, column=col_start).value or "").strip()

        if first_cell and current:
            # Flush previous entry
            entries.append((current_title, "; ".join(current) if current else text))
            current = [text]
        elif first_cell and not current:
            current = [text]
        else:
            current.append(text)

    # Flush last
    if current:
        entries.append((current_title, "; ".join(current)))

    return entries


def format_section(name: str, entries: list) -> str:
    """Format a section as markdown."""
    if not entries:
        return ""

    lines = [f"## {name}\n"]
    for title, body in entries:
        if body:
            # Clean up the text
            body = body.strip()
            if body:
                lines.append(f"- **{title}**：{body}")
    lines.append("")
    return "\n".join(lines)


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else r"C:\Users\KnightCode\Downloads\工具人表26.5 (2).xlsx"
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Sheet1']

    sections = [
        ("贸易站-通用乘区", 2, 4),
        ("但书链", 5, 7),
        ("可露希尔", 8, 10),
        ("制造-赤金散件", 11, 13),
        ("自动化组配置", 14, 16),
        ("红云组", 17, 17),
        ("经验制造", 18, 19),
        ("标准化与通用", 20, 21),
    ]

    out = ["# 公孙长乐工具人表（2026年5月版）\n"]
    out.append("> 来源：@公孙长乐制表，基建指南针系列视频\n")

    for name, cs, ce in sections:
        entries = parse_section(ws, name, cs, ce)
        out.append(format_section(name, entries))

    # Append sheet2 (排班表)
    ws2 = wb['Sheet2']
    out.append("## 三班倒换班原理\n")
    for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row, values_only=True):
        vals = [str(c).strip() for c in row if c is not None and str(c).strip()]
        if vals:
            out.append(" | ".join(vals))

    markdown = "\n".join(out)

    out_path = sys.argv[2] if len(sys.argv) > 2 else "工具人表_RAG.md"
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(markdown)

    print(f"Output: {out_path} ({len(markdown)} chars, {markdown.count(chr(10))} lines)")


if __name__ == "__main__":
    main()
